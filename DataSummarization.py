# -*- coding: utf-8 -*-
"""
Created on Thu Jun 21 22:14:21 2017

@author: Cátia Botas
"""

import xlsxwriter
import openpyxl
import xlrd
import unicodecsv
import statistics
from statistics import StatisticsError
import itertools as its
from collections import Counter
from scipy.stats import linregress
import math

def convertXLS2CSV(ExcelFile, SheetName, CSVFile):
    # Converts an Excel file to a CSV file.
    # If the excel file has multiple worksheets, only the first worksheet is converted.
    # Uses unicodecsv, so it will handle Unicode characters.
    # Uses a recent version of xlrd, so it should handle old .xls and new .xlsx equally well.
    

    wb = xlrd.open_workbook(ExcelFile)
    sh = wb.sheet_by_name(SheetName)

    fh = open(CSVFile,"wb")
    csv_out = unicodecsv.writer(fh, encoding='utf-8')

    for row_number in range (sh.nrows):
        csv_out.writerow(sh.row_values(row_number))

    fh.close()

def ilen(it):
    '''Return the length of an iterable.
    
    >>> ilen(range(7))
    7
    '''
    return sum(1 for _ in it)

def runlength_enc(xs):
    '''Return a run-length encoded version of the stream, xs.
    
    The resulting stream consists of (count, x) pairs.
    
    >>> ys = runlength_enc('AAABBCCC')
    >>> next(ys)
    (3, 'A')
    >>> list(ys)
    [(2, 'B'), (3, 'C')]
    '''
    return ((ilen(gp), x) for x, gp in its.groupby(xs))

def calculateFD(data):
    var = []
    aux = 0 
    for r in range(1, len(data)):
        Nd = len(data) - r
        for i in range(len(data)):
            if (i+r) >= len(data):
                continue
            else:
                aux += (data[i]-data[i+r])**2
        aux = aux/2*Nd
        var.append(aux)
    if 0 in var:
        return 0
    if not var:
        return 0
    x = [math.log10(i) for i in range(1,len(data))]
    y = [math.log10(i) for i in var]
    s = linregress(x,y)
    slope = s[0]
    if math.isnan(slope):
        return 0
    return (4-slope)/2

def calculateSRE(RealData):
    data = [abs(number) for number in RealData]
    ys = runlength_enc(data)
    
    Run = list(ys)
    MaxRun = 0
    for r in range(len(Run)):
        if Run[r][0] > MaxRun:
            MaxRun = Run[r][0]
    
    a = MaxRun
    if max(data) == 0:
        b = range(1)
    else:
        b = range(max(data))
    
    Matrix = [ [0] * a for _ in b]
    
    for i in range(len(Run)):
        runCode = int(float(Run[i][0]))
        Num = int(float(Run[i][1]))
        if Num <= 0:
            Matrix[0][runCode - 1] = Matrix[0][runCode - 1] + 1
        else:
            Matrix[Num - 1][runCode - 1] = Matrix[Num - 1][runCode - 1] + 1
    
    sre = 0
    for r in b:
        for i in range(a):
            sre += Matrix[r][i]/(i+1)**2
                         
    return sre/(sum(sum(Matrix,[])))

def calculateLRE(RealData):
    data = [abs(number) for number in RealData]
    ys = runlength_enc(data)
    
    Run = list(ys)
    MaxRun = 0
    for r in range(len(Run)):
        if Run[r][0] > MaxRun:
            MaxRun = Run[r][0]
    
    a = MaxRun
    if max(data) == 0:
        b = range(1)
    else:
        b = range(max(data))
    
    Matrix = [ [0] * a for _ in b]
    
    for i in range(len(Run)):
        runCode = int(float(Run[i][0]))
        Num = int(float(Run[i][1]))
        if Num <= 0:
            Matrix[0][runCode - 1] = Matrix[0][runCode - 1] + 1
        else:
            Matrix[Num - 1][runCode - 1] = Matrix[Num - 1][runCode - 1] + 1
    
    sre = 0
    for r in b:
        for i in range(a):
            sre += Matrix[r][i]*(i+1)**2
                         
    return sre/(sum(sum(Matrix,[])))

def calculateSK(data, m, var):
    if var == 0:
        return 0
    labels, values = zip(*Counter(data).items())
    
    if len(data) == len(labels):
        return sum((values[i]-m)**3/(len(data)*var**3) for i in range(len(data)))
    else:
        return sum((values[i]-m)**3/(len(data)*var**3) for i in range(len(values)))

def calculateKUR(data, m, var):
    if var == 0:
        return 0
    labels, values = zip(*Counter(data).items())
    
    if len(data) == len(labels):
        return sum((values[i]-m)**4/(len(data)*var**4) for i in range(len(data)))
    else:
        for r in range(min(data), max(data)+1):
            if r not in labels:
                values = list(values)
                values.append(0)
        return sum((values[i]-m)**4/(len(data)*var**4) for i in range(len(values)))
    
def DataSummarization(T0, T3, T6, T12, T18, T24, row, realCol, atualCol, lengthTS, worksheet):
    NominalCol = [39,40,41,42]
    NotIntegerCol = [9,10,11,12,13,14,15,36,38]
    
    if lengthTS == "0m":
        if (T0 is None) or (T0 is "") or (T0 is ' '):
            if realCol in NominalCol:
                atualCol = atualCol + 1
            elif realCol in NotIntegerCol:
                atualCol = atualCol + 4
            else:
                atualCol = atualCol + 9
            return atualCol
        else:
            data = [T0]
            if realCol in NominalCol:
                try:
                    mode = statistics.mode(data)
                except StatisticsError:
                    mode = ""
                worksheet.write(row, atualCol, mode)
                atualCol = atualCol + 1
            elif realCol in NotIntegerCol:
                m = statistics.mean(data)
                try:
                    var = statistics.variance(data,m)
                except StatisticsError:
                    var = 0
                try:
                    mode = round(statistics.mode(data),1)
                except StatisticsError:
                    mode = ""
                worksheet.write(row, atualCol, round(m,1))
                worksheet.write(row, atualCol + 1, round(statistics.median(data),1))
                worksheet.write(row, atualCol + 2, mode)
                worksheet.write(row, atualCol + 3, round(var,1))
                atualCol = atualCol + 4
            else:
                m = statistics.mean(data)
                try:
                    var = statistics.variance(data,m)
                except StatisticsError:
                    var = 0
                try:
                    mode = round(statistics.mode(data),0)
                except StatisticsError:
                    mode = ""
                worksheet.write(row, atualCol, round(m,0))
                worksheet.write(row, atualCol + 1, round(statistics.median(data),0))
                worksheet.write(row, atualCol + 2, mode)
                worksheet.write(row, atualCol + 3, round(var,0))
                worksheet.write(row, atualCol + 4, round(calculateFD(data),0))
                worksheet.write(row, atualCol + 5, round(calculateSRE(data),0))
                worksheet.write(row, atualCol + 6, round(calculateLRE(data),0))
                worksheet.write(row, atualCol + 7, round(calculateSK(data, m, var),0))
                worksheet.write(row, atualCol + 8, round(calculateKUR(data, m, var),0))
                atualCol = atualCol + 9
    elif lengthTS == "3m":
        if (T0 is None) or (T0 is "") or (T0 is ' ') or (T3 is None) or (T3 is "") or (T3 is ' '):
            count = 0
            VerificationT0 = (T0 is None) or (T0 is "") or (T0 is ' ')
            VerificationT3 = (T3 is None) or (T3 is "") or (T3 is ' ')
            if VerificationT0:
                count = count + 1 
            if VerificationT3:
                count = count + 1         
            
            if count == 2:
                if realCol in NominalCol:
                    atualCol = atualCol + 1
                elif realCol in NotIntegerCol:
                    atualCol = atualCol + 4
                else:
                    atualCol = atualCol + 9
                return atualCol
            if count == 1:
                if VerificationT0:
                    data = [T3]
                elif VerificationT3:
                    data = [T0]
        else:
            data = [T0, T3]
            
        #Write in Excel
        if realCol in NominalCol:
            try:
                mode = statistics.mode(data)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, mode)
            atualCol = atualCol + 1
        elif realCol in NotIntegerCol:
            m = statistics.mean(data)
            try:
                var = statistics.variance(data,m)
            except StatisticsError:
                var = 0
            try:
                mode = round(statistics.mode(data),1)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, round(m,1))
            worksheet.write(row, atualCol + 1, round(statistics.median(data),1))
            worksheet.write(row, atualCol + 2, mode)
            worksheet.write(row, atualCol + 3, round(var,1))
            atualCol = atualCol + 4
        else:
            m = statistics.mean(data)
            try:
                var = statistics.variance(data,m)
            except StatisticsError:
                var = 0
            try:
                mode = round(statistics.mode(data),0)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, round(m,0))
            worksheet.write(row, atualCol + 1, round(statistics.median(data),0))
            worksheet.write(row, atualCol + 2, mode)
            worksheet.write(row, atualCol + 3, round(var,0))
            worksheet.write(row, atualCol + 4, round(calculateFD(data),0))
            worksheet.write(row, atualCol + 5, round(calculateSRE(data),0))
            worksheet.write(row, atualCol + 6, round(calculateLRE(data),0))
            worksheet.write(row, atualCol + 7, round(calculateSK(data, m, var),0))
            worksheet.write(row, atualCol + 8, round(calculateKUR(data, m, var),0))
            atualCol = atualCol + 9
    elif lengthTS == "6m":
        if (T0 is None) or (T0 is "") or (T0 is ' ') or (T3 is None) or (T3 is "") or (T3 is ' ') or (T6 is None) or (T6 is "") or (T6 is ' '):
            count = 0
            VerificationT0 = (T0 is None) or (T0 is "") or (T0 is ' ')
            VerificationT3 = (T3 is None) or (T3 is "") or (T3 is ' ')
            VerificationT6 = (T6 is None) or (T6 is "") or (T6 is ' ')
            if VerificationT0:
                count = count + 1 
            if VerificationT3:
                count = count + 1 
            if VerificationT6:
                count = count + 1
            #Don't write anything
            if count == 3:
                if realCol in NominalCol:
                    atualCol = atualCol + 1
                elif realCol in NotIntegerCol:
                    atualCol = atualCol + 4
                else:
                    atualCol = atualCol + 9
                return atualCol
            #When there's 2 missings
            elif count == 2:
                # Atribute values to data
                if VerificationT0 and VerificationT3:
                    data = [T6]
                elif VerificationT0 and VerificationT6:
                    data = [T3]
                else:
                    data = [T0]
            else:
                # Atribute values to data
                if VerificationT0:
                    data = [T3, T6]
                elif VerificationT3:
                    data = [T0, T6]
                else:
                    data = [T0, T3]   
        else:
            # Atribute values to data
            data = [T0, T3, T6]
        #Write in Excel
        if realCol in NominalCol:
            try:
                mode = statistics.mode(data)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, mode)
            atualCol = atualCol + 1
        elif realCol in NotIntegerCol:
            m = statistics.mean(data)
            try:
                var = statistics.variance(data,m)
            except StatisticsError:
                var = 0
            try:
                mode = round(statistics.mode(data),1)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, round(m,1))
            worksheet.write(row, atualCol + 1, round(statistics.median(data),1))
            worksheet.write(row, atualCol + 2, mode)
            worksheet.write(row, atualCol + 3, round(var,1))
            atualCol = atualCol + 4
        else:
            m = statistics.mean(data)
            try:
                var = statistics.variance(data,m)
            except StatisticsError:
                var = 0
            try:
                mode = round(statistics.mode(data),0)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, round(m,0))
            worksheet.write(row, atualCol + 1, round(statistics.median(data),0))
            worksheet.write(row, atualCol + 2, mode)
            worksheet.write(row, atualCol + 3, round(var,0))
            worksheet.write(row, atualCol + 4, round(calculateFD(data),0))
            worksheet.write(row, atualCol + 5, round(calculateSRE(data),0))
            worksheet.write(row, atualCol + 6, round(calculateLRE(data),0))
            worksheet.write(row, atualCol + 7, round(calculateSK(data, m, var),0))
            worksheet.write(row, atualCol + 8, round(calculateKUR(data, m, var),0))
            atualCol = atualCol + 9
            
    elif lengthTS == "12m":
        if (T0 is None) or (T0 is "") or (T0 is ' ') or (T3 is None) or (T3 is "") or (T3 is ' ') or (T6 is None) or (T6 is "") or (T6 is ' ') or (T12 is None) or (T12 is "") or (T12 is ' '):
            count = 0
            VerificationT0 = (T0 is None) or (T0 is "") or (T0 is ' ')
            VerificationT3 = (T3 is None) or (T3 is "") or (T3 is ' ')
            VerificationT6 = (T6 is None) or (T6 is "") or (T6 is ' ')
            VerificationT12 = (T12 is None) or (T12 is "") or (T12 is ' ')
            if VerificationT0:
                count = count + 1 
            if VerificationT3:
                count = count + 1 
            if VerificationT6:
                count = count + 1
            if VerificationT12:
                count = count + 1
            #Don't write anything
            if count == 4:
                if realCol in NominalCol:
                    atualCol = atualCol + 1
                elif realCol in NotIntegerCol:
                    atualCol = atualCol + 4
                else:
                    atualCol = atualCol + 9
                return atualCol
            #When there is 3 missings
            elif count == 3:
                # Atribute values to data
                if VerificationT0 and VerificationT3 and VerificationT6:
                    data = [T12]
                if VerificationT0 and VerificationT3  and VerificationT12:
                    data = [T6]
                if VerificationT0 and VerificationT6 and VerificationT12:
                    data = [T3]
                if VerificationT3 and VerificationT6 and VerificationT12:
                    data = [T0]
            #When there is 3 missings
            elif count == 2:
                # Atribute values to data
                if VerificationT0 and VerificationT3:
                    data = [T6, T12]
                if VerificationT0 and VerificationT6:
                    data = [T3, T12]
                if VerificationT0 and VerificationT12:
                    data = [T3, T6]
                if VerificationT3 and VerificationT6:
                    data = [T0, T12]
                if VerificationT3 and VerificationT12:
                    data = [T0, T6]
                if VerificationT6 and VerificationT12:
                    data = [T0, T3]
            #When there is 1 missing
            elif count == 1:
                # Atribute values to data
                if VerificationT0:
                    data = [T3, T6, T12]
                if VerificationT3:
                    data = [T0, T6, T12]
                if VerificationT6:
                    data = [T0, T3, T12]
                if VerificationT12:
                    data = [T0, T3, T6]
        else:
            # Atribute values to data
            data = [T0, T3, T6, T12]
        #Write in Excel
        if realCol in NominalCol:
            try:
                mode = statistics.mode(data)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, mode)
            atualCol = atualCol + 1
        elif realCol in NotIntegerCol:
            m = statistics.mean(data)
            try:
                var = statistics.variance(data,m)
            except StatisticsError:
                var = 0
            try:
                mode = round(statistics.mode(data),1)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, round(m,1))
            worksheet.write(row, atualCol + 1, round(statistics.median(data),1))
            worksheet.write(row, atualCol + 2, mode)
            worksheet.write(row, atualCol + 3, round(var,1))
            atualCol = atualCol + 4
        else:
            m = statistics.mean(data)
            try:
                var = statistics.variance(data,m)
            except StatisticsError:
                var = 0
            try:
                mode = round(statistics.mode(data),0)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, round(m,0))
            worksheet.write(row, atualCol + 1, round(statistics.median(data),0))
            worksheet.write(row, atualCol + 2, mode)
            worksheet.write(row, atualCol + 3, round(var,0))
            worksheet.write(row, atualCol + 4, round(calculateFD(data),0))
            worksheet.write(row, atualCol + 5, round(calculateSRE(data),0))
            worksheet.write(row, atualCol + 6, round(calculateLRE(data),0))
            worksheet.write(row, atualCol + 7, round(calculateSK(data, m, var),0))
            worksheet.write(row, atualCol + 8, round(calculateKUR(data, m, var),0))
            atualCol = atualCol + 9
    elif lengthTS == "18m":
        if (T0 is None) or (T0 is "") or (T0 is ' ') or (T3 is None) or (T3 is "") or (T3 is ' ') or (T6 is None) or (T6 is "") or (T6 is ' ') or (T12 is None) or (T12 is "") or (T12 is ' ') or (T18 is None) or (T18 is "") or (T18 is ' '):
            VerificationT0 = (T0 is None) or (T0 is "") or (T0 is ' ')
            VerificationT3 = (T3 is None) or (T3 is "") or (T3 is ' ')
            VerificationT6 = (T6 is None) or (T6 is "") or (T6 is ' ')
            VerificationT12 = (T12 is None) or (T12 is "") or (T12 is ' ')
            VerificationT18 = (T18 is None) or (T18 is "") or (T18 is ' ')
            count = 0
            if VerificationT0:
                count = count + 1 
            if VerificationT3:
                count = count + 1 
            if VerificationT6:
                count = count + 1
            if VerificationT12:
                count = count + 1
            if VerificationT18:
                count = count + 1
            #Don't write anything
            if count == 5:
                if realCol in NominalCol:
                    atualCol = atualCol + 1
                elif realCol in NotIntegerCol:
                    atualCol = atualCol + 4
                else:
                    atualCol = atualCol + 9
                return atualCol
            #When there is 4 missings
            if count == 4: 
                if VerificationT0 and VerificationT3 and VerificationT6 and VerificationT12:
                    data = [T18]
                if VerificationT0 and VerificationT3 and VerificationT6 and VerificationT18:
                    data = [T12]
                if VerificationT0 and VerificationT3 and VerificationT12 and VerificationT18:
                    data = [T6]
                if VerificationT0 and VerificationT6 and VerificationT12 and VerificationT18:
                    data = [T3]
                if VerificationT3 and VerificationT6 and VerificationT12 and VerificationT18:
                    data = [T0]
            #When there is 3 missings                    
            if count == 3: 
                if VerificationT0 and VerificationT3 and VerificationT6:
                    data = [T12, T18]
                if VerificationT0 and VerificationT3 and VerificationT12:
                    data = [T6, T18]
                if VerificationT0 and VerificationT3 and VerificationT18:
                    data = [T6, T12]
                if VerificationT0 and VerificationT6 and VerificationT12:
                    data = [T3, T18]
                if VerificationT0 and VerificationT6 and VerificationT18:
                    data = [T3, T12]
                if VerificationT0 and VerificationT12 and VerificationT18:
                    data = [T3, T6]
                if VerificationT3 and VerificationT6 and VerificationT12:
                    data = [T0, T18]
                if VerificationT3 and VerificationT6 and VerificationT18:
                    data = [T0, T12]
                if VerificationT3 and VerificationT12 and VerificationT18:
                    data = [T0, T6]
                if VerificationT6 and VerificationT12 and VerificationT18:
                    data = [T0, T3]
            #When there is 2 missings
            if count == 2:
                if VerificationT0 and VerificationT3:
                    data = [T6, T12, T18]
                if VerificationT0 and VerificationT6:
                    data = [T3, T12, T18]
                if VerificationT0 and VerificationT12:
                    data = [T3, T6, T18]
                if VerificationT0 and VerificationT18:
                    data = [T3, T6, T12]
                if VerificationT3 and VerificationT6:
                    data = [T0, T12, T18]
                if VerificationT3 and VerificationT12:
                    data = [T0, T6, T18]
                if VerificationT3 and VerificationT18:
                    data = [T0, T6, T12]
                if VerificationT6 and VerificationT12:
                    data = [T0, T3, T18]
                if VerificationT6 and VerificationT18:
                    data = [T0, T3, T12]
                if VerificationT12 and VerificationT18:
                    data = [T0, T3, T6]
            #When there is 1 missing
            if count == 1:
                if VerificationT0:
                    data = [T3, T6, T12, T18]
                if VerificationT3:
                    data = [T0, T6, T12, T18]
                if VerificationT6:
                    data = [T0, T3, T12, T18]
                if VerificationT12:
                    data = [T0, T3, T6, T18]
                if VerificationT18:
                    data = [T0, T3, T6, T12]
        else:
            data = [T0, T3, T6, T12, T18]
        #Write in Excel
        if realCol in NominalCol:
            try:
                mode = statistics.mode(data)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, mode)
            atualCol = atualCol + 1
        elif realCol in NotIntegerCol:
            m = statistics.mean(data)
            try:
                var = statistics.variance(data,m)
            except StatisticsError:
                var = 0
            try:
                mode = round(statistics.mode(data),1)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, round(m,1))
            worksheet.write(row, atualCol + 1, round(statistics.median(data),1))
            worksheet.write(row, atualCol + 2, mode)
            worksheet.write(row, atualCol + 3, round(var,1))
            atualCol = atualCol + 4
        else:
            m = statistics.mean(data)
            try:
                var = statistics.variance(data,m)
            except StatisticsError:
                var = 0
            try:
                mode = round(statistics.mode(data),0)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, round(m,0))
            worksheet.write(row, atualCol + 1, round(statistics.median(data),0))
            worksheet.write(row, atualCol + 2, mode)
            worksheet.write(row, atualCol + 3, round(var,0))
            worksheet.write(row, atualCol + 4, round(calculateFD(data),0))
            worksheet.write(row, atualCol + 5, round(calculateSRE(data),0))
            worksheet.write(row, atualCol + 6, round(calculateLRE(data),0))
            worksheet.write(row, atualCol + 7, round(calculateSK(data, m, var),0))
            worksheet.write(row, atualCol + 8, round(calculateKUR(data, m, var),0))
            atualCol = atualCol + 9
    else:
        if (T0 is None) or (T0 is "") or (T0 is ' ') or (T3 is None) or (T3 is "") or (T3 is ' ') or (T6 is None) or (T6 is "") or (T6 is ' ') or (T12 is None) or (T12 is "") or (T12 is ' ') or (T18 is None) or (T18 is "") or (T18 is ' ') or (T24 is None) or (T24 is "") or (T24 is ' '):
            count = 0
            VerificationT0 = (T0 is None) or (T0 is "") or (T0 is ' ')
            VerificationT3 = (T3 is None) or (T3 is "") or (T3 is ' ')
            VerificationT6 = (T6 is None) or (T6 is "") or (T6 is ' ')
            VerificationT12 = (T12 is None) or (T12 is "") or (T12 is ' ')
            VerificationT18 = (T18 is None) or (T18 is "") or (T18 is ' ')
            VerificationT24 = (T24 is None) or (T24 is "") or (T24 is ' ')
            if VerificationT0:
                count = count + 1 
            if VerificationT3:
                count = count + 1 
            if VerificationT6:
                count = count + 1
            if VerificationT12:
                count = count + 1
            if VerificationT18:
                count = count + 1
            if VerificationT24:
                count = count + 1
            #Don't write anything
            if count == 6:
                if realCol in NominalCol:
                    atualCol = atualCol + 1
                elif realCol in NotIntegerCol:
                    atualCol = atualCol + 4
                else:
                    atualCol = atualCol + 9
                return atualCol
            #When there is 5 missings
            if count == 5:
                if VerificationT0 and VerificationT3 and VerificationT6 and VerificationT12 and VerificationT18:
                    data = [T24]
                if VerificationT0 and VerificationT3 and VerificationT6 and VerificationT12 and VerificationT24:
                    data = [T18]
                if VerificationT0 and VerificationT3 and VerificationT6 and VerificationT18 and VerificationT24:
                    data = [T12]
                if VerificationT0 and VerificationT3 and VerificationT12 and VerificationT18 and VerificationT24:
                    data = [T6]
                if VerificationT0 and VerificationT6 and VerificationT12 and VerificationT18 and VerificationT24:
                    data = [T3]
                if VerificationT3 and VerificationT6 and VerificationT12 and VerificationT18 and VerificationT24:
                    data = [T0]
            #When there is 4 missings
            if count == 4:
                if VerificationT0 and VerificationT3 and VerificationT6 and VerificationT12:
                    data = [T18, T24]
                if VerificationT0 and VerificationT3 and VerificationT6 and VerificationT18:
                    data = [T12, T24]
                if VerificationT0 and VerificationT3 and VerificationT6 and VerificationT24:
                    data = [T12, T18]
                if VerificationT0 and VerificationT3 and VerificationT12 and VerificationT18:
                    data = [T6, T24]
                if VerificationT0 and VerificationT3 and VerificationT12 and VerificationT24:
                    data = [T6, T18]
                if VerificationT0 and VerificationT3 and VerificationT18 and VerificationT24:
                    data = [T6, T12]
                if VerificationT0 and VerificationT6 and VerificationT12 and VerificationT18:
                    data = [T3, T24]
                if VerificationT0 and VerificationT6 and VerificationT12 and VerificationT24:
                    data = [T3, T18]
                if VerificationT0 and VerificationT6 and VerificationT18 and VerificationT24:
                    data = [T3, T12]
                if VerificationT0 and VerificationT12 and VerificationT18 and VerificationT24:
                    data = [T3, T6]
                if VerificationT3 and VerificationT6 and VerificationT12 and VerificationT18:
                    data = [T0, T24]
                if VerificationT3 and VerificationT6 and VerificationT12 and VerificationT24:
                    data = [T0, T18]
                if VerificationT3 and VerificationT6 and VerificationT18 and VerificationT24:
                    data = [T0, T12]
                if VerificationT3 and VerificationT12 and VerificationT18 and VerificationT24:
                    data = [T0, T6]
                if VerificationT6 and VerificationT12 and VerificationT18 and VerificationT24:
                    data = [T0, T3]
            #When there is 3 missings
            if count == 3:
                if VerificationT0 and VerificationT3 and VerificationT6:
                    data = [T12, T18, T24]
                if VerificationT0 and VerificationT3 and VerificationT12:
                    data = [T6, T18, T24]
                if VerificationT0 and VerificationT3 and VerificationT18:
                    data = [T6, T12, T24]
                if VerificationT0 and VerificationT3 and VerificationT24:
                    data = [T6, T12, T18]
                if VerificationT0 and VerificationT6 and VerificationT12:
                    data = [T3, T18, T24]
                if VerificationT0 and VerificationT6 and VerificationT18:
                    data = [T3, T12, T24]
                if VerificationT0 and VerificationT6 and VerificationT24:
                    data = [T3, T12, T18]
                if VerificationT0 and VerificationT12 and VerificationT18:
                    data = [T3, T6, T24]
                if VerificationT0 and VerificationT12 and VerificationT24:
                    data = [T3, T6, T18]
                if VerificationT0 and VerificationT18 and VerificationT24:
                    data = [T3, T6, T12]
                if VerificationT3 and VerificationT6 and VerificationT12:
                    data = [T0, T18, T24]
                if VerificationT3 and VerificationT6 and VerificationT18:
                    data = [T0, T12, T24]
                if VerificationT3 and VerificationT6 and VerificationT24:
                    data = [T0, T12, T18]
                if VerificationT3 and VerificationT12 and VerificationT18:
                    data = [T0, T6, T24]
                if VerificationT3 and VerificationT12 and VerificationT24:
                    data = [T0, T6, T18]
                if VerificationT3 and VerificationT18 and VerificationT24:
                    data = [T0, T6, T12]
                if VerificationT6 and VerificationT12 and VerificationT18:
                    data = [T0, T3, T24]
                if VerificationT6 and VerificationT12 and VerificationT24:
                    data = [T0, T3, T18]
                if VerificationT6 and VerificationT18 and VerificationT24:
                    data = [T0, T3, T12]
                if VerificationT12 and VerificationT18 and VerificationT24:
                    data = [T0, T3, T6]
            #When there is 2 missings
            if count == 2:
                if VerificationT0 and VerificationT3:
                    data = [T6, T12, T18, T24]
                if VerificationT0 and VerificationT6:
                    data = [T3, T12, T18, T24]
                if VerificationT0 and VerificationT12:
                    data = [T3, T6, T18, T24]
                if VerificationT0 and VerificationT18:
                    data = [T3, T6, T12, T24]
                if VerificationT0 and VerificationT24:
                    data = [T3, T6, T12, T18]
                if VerificationT3 and VerificationT6:
                    data = [T0, T12, T18, T24]
                if VerificationT3 and VerificationT12:
                    data = [T0, T6, T18, T24]
                if VerificationT3 and VerificationT18:
                    data = [T0, T6, T12, T24]
                if VerificationT3 and VerificationT24:
                    data = [T0, T6, T12, T18]
                if VerificationT6 and VerificationT12:
                    data = [T0, T3, T18, T24]
                if VerificationT6 and VerificationT18:
                    data = [T0, T3, T12, T24]
                if VerificationT6 and VerificationT24:
                    data = [T0, T3, T12, T18]
                if VerificationT12 and VerificationT18:
                    data = [T0, T3, T6, T24]
                if VerificationT12 and VerificationT24:
                    data = [T0, T3, T6, T18]
                if VerificationT18 and VerificationT24:
                    data = [T0, T3, T6, T12]
            #When there is 1 missing
            if count == 1:
                if VerificationT0:
                    data = [T3, T6, T12, T18, T24]
                if VerificationT3:
                    data = [T0, T6, T12, T18, T24]
                if VerificationT6:
                    data = [T0, T3, T12, T18, T24]
                if VerificationT12:
                    data = [T0, T3, T6, T18, T24]
                if VerificationT18:
                    data = [T0, T3, T6, T12, T24]
                if VerificationT24:
                    data = [T0, T3, T6, T12, T18]
        else:
            data = [T0, T3, T6, T12, T18, T24]
        #Write in Excel
        if realCol in NominalCol:
            try:
                mode = statistics.mode(data)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, mode)
            atualCol = atualCol + 1
        elif realCol in NotIntegerCol:
            m = statistics.mean(data)
            try:
                var = statistics.variance(data,m)
            except StatisticsError:
                var = 0
            try:
                mode = round(statistics.mode(data),1)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, round(m,1))
            worksheet.write(row, atualCol + 1, round(statistics.median(data),1))
            worksheet.write(row, atualCol + 2, mode)
            worksheet.write(row, atualCol + 3, round(var,1))
            atualCol = atualCol + 4
        else:
            m = statistics.mean(data)
            try:
                var = statistics.variance(data,m)
            except StatisticsError:
                var = 0
            try:
                mode = round(statistics.mode(data),0)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, round(m,0))
            worksheet.write(row, atualCol + 1, round(statistics.median(data),0))
            worksheet.write(row, atualCol + 2, mode)
            worksheet.write(row, atualCol + 3, round(var,0))
            worksheet.write(row, atualCol + 4, round(calculateFD(data),0))
            worksheet.write(row, atualCol + 5, round(calculateSRE(data),0))
            worksheet.write(row, atualCol + 6, round(calculateLRE(data),0))
            worksheet.write(row, atualCol + 7, round(calculateSK(data, m, var),0))
            worksheet.write(row, atualCol + 8, round(calculateKUR(data, m, var),0))
            atualCol = atualCol + 9
    return atualCol


def WriteInformation (worksheet, MaxRow, MaxColumn, sheet0):
    
    NominalCol = [39,40,41,42]
    NotIntegerCol = [9,10,11,12,13,14,15,36,38]
    
    # Write pacients information
    for i in range(2,4):
        worksheet.write(0, i-2, sheet0.cell(row = 1, column = i + 1).value)
    
    #Write features summarization 
    indexCol = 2    
    for r in range(4, MaxColumn):
        if r == MaxColumn - 1:
            worksheet.write(0,indexCol, sheet0.cell(row = 1, column = r + 1).value)
        elif r in NominalCol:
            worksheet.write(0,indexCol, sheet0.cell(row = 1, column = r + 1).value + "_MODE")
            indexCol = indexCol + 1
        elif r in NotIntegerCol:
            worksheet.write(0, indexCol, sheet0.cell(row = 1, column = r + 1).value + "_MEAN")
            worksheet.write(0, indexCol + 1, sheet0.cell(row = 1, column = r + 1).value + "_MEDIAN")
            worksheet.write(0, indexCol + 2, sheet0.cell(row = 1, column = r + 1).value + "_MODE")
            worksheet.write(0, indexCol + 3, sheet0.cell(row = 1, column = r + 1).value + "_VAR")
            indexCol = indexCol + 4
        else:
            worksheet.write(0, indexCol, sheet0.cell(row = 1, column = r + 1).value + "_MEAN")
            worksheet.write(0, indexCol + 1, sheet0.cell(row = 1, column = r + 1).value + "_MEDIAN")
            worksheet.write(0, indexCol + 2, sheet0.cell(row = 1, column = r + 1).value + "_MODE")
            worksheet.write(0, indexCol + 3, sheet0.cell(row = 1, column = r + 1).value + "_VAR")
            worksheet.write(0, indexCol + 4, sheet0.cell(row = 1, column = r + 1).value + "_FD")
            worksheet.write(0, indexCol + 5, sheet0.cell(row = 1, column = r + 1).value + "_SRE")
            worksheet.write(0, indexCol + 6, sheet0.cell(row = 1, column = r + 1).value + "_LRE")
            worksheet.write(0, indexCol + 7, sheet0.cell(row = 1, column = r + 1).value + "_SK")
            worksheet.write(0, indexCol + 8, sheet0.cell(row = 1, column = r + 1).value + "_KUR")
            indexCol = indexCol + 9

def main():
    
    dataFileName = 'new_Reuma_ImputatedMissingValuesSTATS.xlsx'
    #dataFileName = 'new_reuma.xlsx'
    
    # To read from new_reuma.xlsx file
    wb = openpyxl.load_workbook(dataFileName, data_only = True)
    estaticas = wb.get_sheet_by_name('Estaticas')
    sheet0 = wb.get_sheet_by_name('0 meses')
    sheet3 = wb.get_sheet_by_name('3 meses')
    sheet6 = wb.get_sheet_by_name('6 meses')
    sheet12 = wb.get_sheet_by_name('12 meses')
    sheet18 = wb.get_sheet_by_name('18 meses')
    sheet24 = wb.get_sheet_by_name('24 meses')
    
    # ==================================================================
    # To write in datasummarization.xlsx file
    workbook = xlsxwriter.Workbook('new_reuma_DataSummarization.xlsx')
    worksheet = workbook.add_worksheet('Summarization')
    
    MaxColumn = sheet0.max_column
    MaxRow = sheet0.max_row
    
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
        
    # ==================================================
    # NEW_REUMA --> classFile
    for r in range(1, MaxRow):
        atualCol = 2
        for k in range(4, MaxColumn):
            if k == MaxColumn - 1:
                if estaticas.cell(row = r + 1, column = 4).value == '0m':
                    worksheet.write(r, atualCol, sheet0.cell(row = r + 1, column = k + 1).value)
                elif estaticas.cell(row = r + 1, column = 4).value == '3m':
                    worksheet.write(r, atualCol, sheet3.cell(row = r + 1, column = k + 1).value)
                elif estaticas.cell(row = r + 1, column = 4).value == '6m':
                    worksheet.write(r, atualCol, sheet6.cell(row = r + 1, column = k + 1).value)
                elif estaticas.cell(row = r + 1, column = 4).value == '12m':
                    worksheet.write(r, atualCol, sheet12.cell(row = r + 1, column = k + 1).value)
                elif estaticas.cell(row = r + 1, column = 4).value == '18m':
                    worksheet.write(r, atualCol, sheet18.cell(row = r + 1, column = k + 1).value)
                else:
                    worksheet.write(r, atualCol, sheet24.cell(row = r + 1, column = k + 1).value)
            else:
                atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                r, k, atualCol, estaticas.cell(row = r + 1 , column = 4).value, worksheet)
                worksheet.write(r, 0, sheet0.cell(row = r + 1, column = 3).value)
                worksheet.write(r, 1, sheet0.cell(row = r + 1, column = 4).value)
    
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = 'new_reuma_DataSummarization.xlsx'
    SheetName = 'Summarization'
    CSVFile = 'new_reuma_DataSummarization.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT3_classLabelT3
    NewFileName = 'classFileSummarization_T0dataT3_classLabelT3.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet3.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet3.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '3m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT3_classLabelT3.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT3_classLabelT6
    NewFileName = 'classFileSummarization_T0dataT3_classLabelT6.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet6.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet6.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '3m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT3_classLabelT6.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT3_classLabelT12
    NewFileName = 'classFileSummarization_T0dataT3_classLabelT12.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet12.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet12.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '3m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT3_classLabelT12.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT3_classLabelT18
    NewFileName = 'classFileSummarization_T0dataT3_classLabelT18.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet18.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet18.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '3m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT3_classLabelT18.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT3_classLabelT24
    NewFileName = 'classFileSummarization_T0dataT3_classLabelT24.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet24.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet24.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '3m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT3_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT6_classLabelT6
    NewFileName = 'classFileSummarization_T0dataT6_classLabelT6.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet6.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet6.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '6m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT6_classLabelT6.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT6_classLabelT12
    NewFileName = 'classFileSummarization_T0dataT6_classLabelT12.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet12.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet12.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '6m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT6_classLabelT12.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT6_classLabelT18
    NewFileName = 'classFileSummarization_T0dataT6_classLabelT18.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet18.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet18.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '6m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT6_classLabelT18.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT6_classLabelT24
    NewFileName = 'classFileSummarization_T0dataT6_classLabelT24.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet24.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet24.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '6m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT6_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT12_classLabelT12
    NewFileName = 'classFileSummarization_T0dataT12_classLabelT12.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet12.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet12.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '12m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT12_classLabelT12.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT12classLabelT18
    NewFileName = 'classFileSummarization_T0dataT12_classLabelT18.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet18.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet18.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '12m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT12_classLabelT18.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT12_classLabelT24
    NewFileName = 'classFileSummarization_T0dataT12_classLabelT24.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet24.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet24.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '12m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT12_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT18_classLabelT18
    NewFileName = 'classFileSummarization_T0dataT18_classLabelT18.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet18.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet18.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '18m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT18_classLabelT18.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT18_classLabelT24
    NewFileName = 'classFileSummarization_T0dataT18_classLabelT24.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet24.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet24.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '18m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT18_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT24_classLabelT24
    NewFileName = 'classFileSummarization_T0dataT24_classLabelT24.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Summarization')
    WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet24.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet24.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataSummarization(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '24m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Summarization'
    CSVFile = 'classFileSummarization_T0dataT24_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
   
if __name__ == "__main__":
    main()
