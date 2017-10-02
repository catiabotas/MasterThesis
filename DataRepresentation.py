# -*- coding: utf-8 -*-
"""
Created on Thu Jun 21 22:14:21 2017

@author: Cátia Botas
"""

import xlsxwriter
import openpyxl
import xlrd
import unicodecsv
import statistics
from statistics import StatisticsError
import itertools as its
import numpy as np
import pywt
import pandas as pd

def all_same(items):
    items = np.asarray(items)
    return all(x == items[0] for x in items)

def convertXLS2CSV(ExcelFile, SheetName, CSVFile):
    # Converts an Excel file to a CSV file.
    # If the excel file has multiple worksheets, only the first worksheet is converted.
    # Uses unicodecsv, so it will handle Unicode characters.
    # Uses a recent version of xlrd, so it should handle old .xls and new .xlsx equally well.
    

    wb = xlrd.open_workbook(ExcelFile)
    sh = wb.sheet_by_name(SheetName)

    fh = open(CSVFile,"wb")
    csv_out = unicodecsv.writer(fh, encoding='utf-8')

    for row_number in range (sh.nrows):
        csv_out.writerow(sh.row_values(row_number))

    fh.close()

def ilen(it):
    '''Return the length of an iterable.
    
    >>> ilen(range(7))
    7
    '''
    return sum(1 for _ in it)

def runlength_enc(xs):
    '''Return a run-length encoded version of the stream, xs.
    
    The resulting stream consists of (count, x) pairs.
    
    >>> ys = runlength_enc('AAABBCCC')
    >>> next(ys)
    (3, 'A')
    >>> list(ys)
    [(2, 'B'), (3, 'C')]
    '''
    return ((ilen(gp), x) for x, gp in its.groupby(xs))

def DFT(data):
    return np.fft.rfft(data)
    

def DWT(data):
    output = []
    (cA, cD) = pywt.dwt(data, 'db1')
    output.append(cA)
    """output.append(cD)"""
    return output

def znormalization(ts):
    """
    ts - each column of ts is a time series (np.ndarray)
    """
    mus = ts.mean(axis = 0)
    stds = ts.std(axis = 0)
    return (ts - mus) / stds

def PAA(data, pieces):
    """
    ts: the columns of which are time series represented by e.g. np.array
    n_pieces: M equally sized piecies into which the original ts is splitted
    """
    splitted = np.array_split(data, pieces) ## along columns as we want
    return np.asarray(list(map(lambda xs: xs.mean(axis = 0), splitted)))

def SDL(data):
    n = len(data)
    output = []
    for i in range(n - 1):
        if data[i] < data [i + 1]:
            output.append('up')
        elif data[i] > data [i + 1]:
            output.append('down')
        else:
            output.append('stable')
    return output

def SAX(data, pieces, alphabet):
    """
    ts: columns of which are time serieses represented by np.array
    n_pieces: number of segments in paa transformation
    alphabet: the letters to be translated to, e.g. "abcd", "ab"
    return np.array of ts's sax transformation
    Steps:
    1. znormalize
    2. ppa
    3. find norm distribution breakpoints by scipy.stats
    4. convert ppa transformation into strings
    """
    from scipy.stats import norm
    alphabet_sz = len(alphabet)
    thrholds = norm.ppf(np.linspace(1./alphabet_sz, 
                                    1-1./alphabet_sz, 
                                    alphabet_sz-1))
    def translate(ts_values):
        return np.asarray([(alphabet[0] if ts_value < thrholds[0]
                else (alphabet[-1] if ts_value > thrholds[-1]
                      else alphabet[np.where(thrholds <= ts_value)[0][-1]+1]))
                           for ts_value in ts_values])
    paa_ts = PAA(znormalization(data), pieces)
    return np.apply_along_axis(translate, 0, paa_ts)

"""def APCA(data):
    
def PLA(data):
"""   
def clipping(data):
    n = len(data)
    mean = statistics.mean(data)
    output = []
    for i in range(n):
        if data[i] > mean:
            output.append(1)
        else:
            output.append(0)
    return output

def DataRepresentation(T0, T3, T6, T12, T18, T24, row, realCol, atualCol, lengthTS, worksheet):
    NominalCol = [39,40,41,42]
    """NotIntegerCol = [9,10,11,12,13,14,15,36,38]"""
    
    if lengthTS == "0m":
        return atualCol + 1
    elif lengthTS == "3m":
        if (T0 is None) or (T0 is "") or (T0 is ' ') or (T3 is None) or (T3 is "") or (T3 is ' '):
            count = 0
            VerificationT0 = (T0 is None) or (T0 is "") or (T0 is ' ')
            VerificationT3 = (T3 is None) or (T3 is "") or (T3 is ' ')
            if VerificationT0:
                count = count + 1 
            if VerificationT3:
                count = count + 1         
            if (count == 2) or (count == 1):
                if realCol in NominalCol:
                    atualCol = atualCol + 1
                else:
                    atualCol = atualCol + 10
                return atualCol
        else:
            data = [T0, T3]
            
        #Write in Excel
        if realCol in NominalCol:
            try:
                mode = statistics.mode(data)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, mode)
            atualCol = atualCol + 1
        else:
            dft = DFT(data)
            dft = dft.tolist()
            none = [""]
            if len(dft) == 1:
                dft = dft + none
            dwt = DWT(data)
            data = np.asarray(data)
            data_ = pd.DataFrame({"ts1": data})
            paa = PAA (data_, 2)
            sdl = SDL(data)
            if all_same(data_):
                sax_array = ['a','a']
            else:
                sax = SAX(data_, 2, "abc")
                sax_array = []
                for k in range(len(sax)):
                    sax_array.append(sax[k,0])
            clip = clipping(data)
            dwt_array = []
            paa_array = []
            for i in range(len(dwt)):
                dwt_array.append(dwt[i].tolist())
            dwt_array = dwt_array[0]
            for l in range(len(paa)):
                paa_array.append(paa[l].tolist())
            finaldata = dft + dwt_array + paa_array + sdl + sax_array + clip
            for k in range(len(finaldata)): 
                worksheet.write(row, atualCol + k, str(finaldata[k]))
            atualCol = atualCol + k + 1
    elif lengthTS == "6m":
        if (T0 is None) or (T0 is "") or (T0 is ' ') or (T3 is None) or (T3 is "") or (T3 is ' ') or (T6 is None) or (T6 is "") or (T6 is ' '):
            count = 0
            VerificationT0 = (T0 is None) or (T0 is "") or (T0 is ' ')
            VerificationT3 = (T3 is None) or (T3 is "") or (T3 is ' ')
            VerificationT6 = (T6 is None) or (T6 is "") or (T6 is ' ')
            if VerificationT0:
                count = count + 1 
            if VerificationT3:
                count = count + 1 
            if VerificationT6:
                count = count + 1
            #Don't write anything
            if (count == 3) or (count == 2):
                if realCol in NominalCol:
                    atualCol = atualCol + 1
                else:
                    atualCol = atualCol + 15
                return atualCol
            #When there's 1 missing
            else:
                # Atribute values to data
                if VerificationT0:
                    data = [T3, T6]
                elif VerificationT3:
                    data = [T0, T6]
                else:
                    data = [T0, T3]   
        else:
            # Atribute values to data
            data = [T0, T3, T6]
        #Write in Excel
        if realCol in NominalCol:
            try:
                mode = statistics.mode(data)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, mode)
            atualCol = atualCol + 1
        else:
            dft = DFT(data)
            dft = dft.tolist()
            dwt = DWT(data)
            none = [""]
            paa_array = []
            r = len(dft)
            if r == 1:
                dft = dft + none
            data = np.asarray(data)
            data_ = pd.DataFrame({"ts1": data})
            size_data = len(data)
            if size_data == 3:
                paa = PAA(data_, 3)
                for l in range(len(paa)):
                    paa_array.append(paa[l].tolist())
                if all_same(data_):
                    sax_array = ['a','a','a']
                else:
                    sax = SAX(data_, 3, "abc")
                    sax_array = []
                    for k in range(len(sax)):
                        sax_array.append(sax[k,0])
                sdl = SDL(data)
                clip = clipping(data)
            elif size_data == 2:
                paa = PAA(data_, 2)
                for l in range(len(paa)):
                    paa_array.append(paa[l].tolist())
                paa_array = paa_array + none 
                if all_same(data_):
                    sax_array = ['a','a']
                else:
                    sax = SAX(data_, 2, "abc")
                    sax_array = []
                    for k in range(len(sax)):
                        sax_array.append(sax[k,0])
                sax_array = sax_array + none
                sdl = SDL(data)
                sdl = sdl + none
                clip = clipping(data)
                clip = clip + none
            dwt_array = []
            for i in range(len(dwt)):
                dwt_array.append(dwt[i].tolist())
            dwt_array = dwt_array[0]
            if len(dwt_array) == 1:
                dwt_array = dwt_array + none
            finaldata = dft + dwt_array + paa_array + sdl + sax_array + clip
            for k in range(len(finaldata)): 
                worksheet.write(row, atualCol + k, str(finaldata[k]))
            atualCol = atualCol + k + 1      
    elif lengthTS == "12m":
        if (T0 is None) or (T0 is "") or (T0 is ' ') or (T3 is None) or (T3 is "") or (T3 is ' ') or (T6 is None) or (T6 is "") or (T6 is ' ') or (T12 is None) or (T12 is "") or (T12 is ' '):
            count = 0
            VerificationT0 = (T0 is None) or (T0 is "") or (T0 is ' ')
            VerificationT3 = (T3 is None) or (T3 is "") or (T3 is ' ')
            VerificationT6 = (T6 is None) or (T6 is "") or (T6 is ' ')
            VerificationT12 = (T12 is None) or (T12 is "") or (T12 is ' ')
            if VerificationT0:
                count = count + 1 
            if VerificationT3:
                count = count + 1 
            if VerificationT6:
                count = count + 1
            if VerificationT12:
                count = count + 1
            #Don't write anything
            if (count == 4) or (count == 3):
                if realCol in NominalCol:
                    atualCol = atualCol + 1
                else:
                    atualCol = atualCol + 20
                return atualCol
            #When there is 3 missings
            elif count == 2:
                # Atribute values to data
                if VerificationT0 and VerificationT3:
                    data = [T6, T12]
                if VerificationT0 and VerificationT6:
                    data = [T3, T12]
                if VerificationT0 and VerificationT12:
                    data = [T3, T6]
                if VerificationT3 and VerificationT6:
                    data = [T0, T12]
                if VerificationT3 and VerificationT12:
                    data = [T0, T6]
                if VerificationT6 and VerificationT12:
                    data = [T0, T3]
            #When there is 1 missing
            elif count == 1:
                # Atribute values to data
                if VerificationT0:
                    data = [T3, T6, T12]
                if VerificationT3:
                    data = [T0, T6, T12]
                if VerificationT6:
                    data = [T0, T3, T12]
                if VerificationT12:
                    data = [T0, T3, T6]
        else:
            # Atribute values to data
            data = [T0, T3, T6, T12]
        #Write in Excel
        if realCol in NominalCol:
            try:
                mode = statistics.mode(data)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, mode)
            atualCol = atualCol + 1
        else:
            dft = DFT(data)
            dft = dft.tolist()
            dwt = DWT(data)
            none = [""]
            paa_array = []
            r = len(dft)
            if r == 1:
                dft = dft + none + none 
            elif r == 2:
                dft = dft + none 
            data = np.asarray(data)
            data_ = pd.DataFrame({"ts1": data})
            size_data = len(data)
            if size_data == 4:
                paa = PAA (data_, 4)
                for l in range(len(paa)):
                    paa_array.append(paa[l].tolist())
                if all_same(data_):
                    sax_array = ['a','a','a','a']
                else:
                    sax = SAX(data_, 4, "abc")
                    sax_array = []
                    for k in range(len(sax)):
                        sax_array.append(sax[k,0])
                sdl = SDL(data)
                clip = clipping(data)
            elif size_data == 3:
                paa = PAA (data_, 3)
                for l in range(len(paa)):
                    paa_array.append(paa[l].tolist())
                paa_array = paa_array + none
                if all_same(data_):
                    sax_array = ['a','a','a']
                else:
                    sax = SAX(data_, 3, "abc")
                    sax_array = []
                    for k in range(len(sax)):
                        sax_array.append(sax[k,0])
                sax_array = sax_array + none
                sdl = SDL(data)
                sdl = sdl + none
                clip = clipping(data)
                clip = clip + none
            else:
                paa = PAA (data_, 2)
                for l in range(len(paa)):
                    paa_array.append(paa[l].tolist())
                paa_array = paa_array + none + none
                if all_same(data_):
                    sax_array = ['a','a']
                else:
                    sax = SAX(data_, 2, "abc")
                    sax_array = []
                    for k in range(len(sax)):
                        sax_array.append(sax[k,0])
                sax_array = sax_array + none + none
                sdl = SDL(data)
                sdl = sdl + none + none
                clip = clipping(data)
                clip = clip + none + none
            dwt_array = []
            for i in range(len(dwt)):
                dwt_array.append(dwt[i].tolist())
            dwt_array = dwt_array[0]
            r = len(dwt_array)
            if r == 1:
                dwt_array = dwt_array + none
            finaldata = dft + dwt_array + paa_array + sdl + sax_array + clip
            for k in range(len(finaldata)): 
                worksheet.write(row, atualCol + k, str(finaldata[k]))
            atualCol = atualCol + k + 1
    elif lengthTS == "18m":
        if (T0 is None) or (T0 is "") or (T0 is ' ') or (T3 is None) or (T3 is "") or (T3 is ' ') or (T6 is None) or (T6 is "") or (T6 is ' ') or (T12 is None) or (T12 is "") or (T12 is ' ') or (T18 is None) or (T18 is "") or (T18 is ' '):
            VerificationT0 = (T0 is None) or (T0 is "") or (T0 is ' ')
            VerificationT3 = (T3 is None) or (T3 is "") or (T3 is ' ')
            VerificationT6 = (T6 is None) or (T6 is "") or (T6 is ' ')
            VerificationT12 = (T12 is None) or (T12 is "") or (T12 is ' ')
            VerificationT18 = (T18 is None) or (T18 is "") or (T18 is ' ')
            count = 0
            if VerificationT0:
                count = count + 1 
            if VerificationT3:
                count = count + 1 
            if VerificationT6:
                count = count + 1
            if VerificationT12:
                count = count + 1
            if VerificationT18:
                count = count + 1
            #Don't write anything
            if (count == 5) or (count == 4):
                if realCol in NominalCol:
                    atualCol = atualCol + 1
                else:
                    atualCol = atualCol + 25
                return atualCol
            #When there is 3 missings                    
            if count == 3: 
                if VerificationT0 and VerificationT3 and VerificationT6:
                    data = [T12, T18]
                if VerificationT0 and VerificationT3 and VerificationT12:
                    data = [T6, T18]
                if VerificationT0 and VerificationT3 and VerificationT18:
                    data = [T6, T12]
                if VerificationT0 and VerificationT6 and VerificationT12:
                    data = [T3, T18]
                if VerificationT0 and VerificationT6 and VerificationT18:
                    data = [T3, T12]
                if VerificationT0 and VerificationT12 and VerificationT18:
                    data = [T3, T6]
                if VerificationT3 and VerificationT6 and VerificationT12:
                    data = [T0, T18]
                if VerificationT3 and VerificationT6 and VerificationT18:
                    data = [T0, T12]
                if VerificationT3 and VerificationT12 and VerificationT18:
                    data = [T0, T6]
                if VerificationT6 and VerificationT12 and VerificationT18:
                    data = [T0, T3]
            #When there is 2 missings
            if count == 2:
                if VerificationT0 and VerificationT3:
                    data = [T6, T12, T18]
                if VerificationT0 and VerificationT6:
                    data = [T3, T12, T18]
                if VerificationT0 and VerificationT12:
                    data = [T3, T6, T18]
                if VerificationT0 and VerificationT18:
                    data = [T3, T6, T12]
                if VerificationT3 and VerificationT6:
                    data = [T0, T12, T18]
                if VerificationT3 and VerificationT12:
                    data = [T0, T6, T18]
                if VerificationT3 and VerificationT18:
                    data = [T0, T6, T12]
                if VerificationT6 and VerificationT12:
                    data = [T0, T3, T18]
                if VerificationT6 and VerificationT18:
                    data = [T0, T3, T12]
                if VerificationT12 and VerificationT18:
                    data = [T0, T3, T6]
            #When there is 1 missing
            if count == 1:
                if VerificationT0:
                    data = [T3, T6, T12, T18]
                if VerificationT3:
                    data = [T0, T6, T12, T18]
                if VerificationT6:
                    data = [T0, T3, T12, T18]
                if VerificationT12:
                    data = [T0, T3, T6, T18]
                if VerificationT18:
                    data = [T0, T3, T6, T12]
        else:
            data = [T0, T3, T6, T12, T18]
        #Write in Excel
        if realCol in NominalCol:
            try:
                mode = statistics.mode(data)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, mode)
            atualCol = atualCol + 1
        else:
            dft = DFT(data)
            dft = dft.tolist()
            dwt = DWT(data)
            none = [""]
            paa_array = []
            r = len(dft)
            if r == 1:
                dft = dft + none + none  
            elif r == 2:
                dft = dft + none 
            data = np.asarray(data)
            data_ = pd.DataFrame({"ts1": data})
            size_data = len(data)
            if size_data == 5:
                paa = PAA (data_, 5)
                for l in range(len(paa)):
                    paa_array.append(paa[l].tolist())
                if all_same(data_):
                    sax_array = ['a','a','a','a','a']
                else:
                    sax = SAX(data_, 5, "abc")
                    sax_array = []
                    for k in range(len(sax)):
                        sax_array.append(sax[k,0])
                sdl = SDL(data)
                clip = clipping(data)
            elif size_data == 4:
                paa = PAA (data_, 4)
                for l in range(len(paa)):
                    paa_array.append(paa[l].tolist())
                paa_array = paa_array + none
                if all_same(data_):
                    sax_array = ['a','a','a','a']
                else:
                    sax = SAX(data_, 4, "abc")
                    sax_array = []
                    for k in range(len(sax)):
                        sax_array.append(sax[k,0])
                sax_array = sax_array + none
                sdl = SDL(data)
                sdl = sdl + none
                clip = clipping(data)
                clip = clip + none
            elif size_data == 3:
                paa = PAA (data_, 3)
                for l in range(len(paa)):
                    paa_array.append(paa[l].tolist())
                paa_array = paa_array + none + none
                if all_same(data_):
                    sax_array = ['a','a','a']
                else:
                    sax = SAX(data_, 3, "abc")
                    sax_array = []
                    for k in range(len(sax)):
                        sax_array.append(sax[k,0])
                sax_array = sax_array + none + none
                sdl = SDL(data)
                sdl = sdl + none + none
                clip = clipping(data)
                clip = clip + none + none
            else:
                paa = PAA (data_, 2)
                for l in range(len(paa)):
                    paa_array.append(paa[l].tolist())
                paa_array = paa_array + none + none + none
                if all_same(data_):
                    sax_array = ['a','a']
                else:
                    sax = SAX(data_, 2, "abc")
                    sax_array = []
                    for k in range(len(sax)):
                        sax_array.append(sax[k,0])
                sax_array = sax_array + none + none + none
                sdl = SDL(data)
                sdl = sdl + none + none + none
                clip = clipping(data)
                clip = clip + none + none + none
            dwt_array = []
            for i in range(len(dwt)):
                dwt_array.append(dwt[i].tolist())
            dwt_array = dwt_array[0]
            r = len(dwt_array)
            if r == 1:
                dwt_array = dwt_array + none + none
            elif r == 2:
                dwt_array = dwt_array + none
            finaldata = dft + dwt_array + paa_array + sdl + sax_array + clip
            size_finaldata = len(finaldata)
            for k in range(size_finaldata): 
                worksheet.write(row, atualCol + k, str(finaldata[k]))
            atualCol = atualCol + k + 1
    else:
        if (T0 is None) or (T0 is "") or (T0 is ' ') or (T3 is None) or (T3 is "") or (T3 is ' ') or (T6 is None) or (T6 is "") or (T6 is ' ') or (T12 is None) or (T12 is "") or (T12 is ' ') or (T18 is None) or (T18 is "") or (T18 is ' ') or (T24 is None) or (T24 is "") or (T24 is ' '):
            count = 0
            VerificationT0 = (T0 is None) or (T0 is "") or (T0 is ' ')
            VerificationT3 = (T3 is None) or (T3 is "") or (T3 is ' ')
            VerificationT6 = (T6 is None) or (T6 is "") or (T6 is ' ')
            VerificationT12 = (T12 is None) or (T12 is "") or (T12 is ' ')
            VerificationT18 = (T18 is None) or (T18 is "") or (T18 is ' ')
            VerificationT24 = (T24 is None) or (T24 is "") or (T24 is ' ')
            if VerificationT0:
                count = count + 1 
            if VerificationT3:
                count = count + 1 
            if VerificationT6:
                count = count + 1
            if VerificationT12:
                count = count + 1
            if VerificationT18:
                count = count + 1
            if VerificationT24:
                count = count + 1
            #Don't write anything
            if (count == 6) or (count == 5):
                if realCol in NominalCol:
                    atualCol = atualCol + 1
                else:
                    atualCol = atualCol + 30
                return atualCol
            #When there is 4 missings
            if count == 4:
                if VerificationT0 and VerificationT3 and VerificationT6 and VerificationT12:
                    data = [T18, T24]
                if VerificationT0 and VerificationT3 and VerificationT6 and VerificationT18:
                    data = [T12, T24]
                if VerificationT0 and VerificationT3 and VerificationT6 and VerificationT24:
                    data = [T12, T18]
                if VerificationT0 and VerificationT3 and VerificationT12 and VerificationT18:
                    data = [T6, T24]
                if VerificationT0 and VerificationT3 and VerificationT12 and VerificationT24:
                    data = [T6, T18]
                if VerificationT0 and VerificationT3 and VerificationT18 and VerificationT24:
                    data = [T6, T12]
                if VerificationT0 and VerificationT6 and VerificationT12 and VerificationT18:
                    data = [T3, T24]
                if VerificationT0 and VerificationT6 and VerificationT12 and VerificationT24:
                    data = [T3, T18]
                if VerificationT0 and VerificationT6 and VerificationT18 and VerificationT24:
                    data = [T3, T12]
                if VerificationT0 and VerificationT12 and VerificationT18 and VerificationT24:
                    data = [T3, T6]
                if VerificationT3 and VerificationT6 and VerificationT12 and VerificationT18:
                    data = [T0, T24]
                if VerificationT3 and VerificationT6 and VerificationT12 and VerificationT24:
                    data = [T0, T18]
                if VerificationT3 and VerificationT6 and VerificationT18 and VerificationT24:
                    data = [T0, T12]
                if VerificationT3 and VerificationT12 and VerificationT18 and VerificationT24:
                    data = [T0, T6]
                if VerificationT6 and VerificationT12 and VerificationT18 and VerificationT24:
                    data = [T0, T3]
            #When there is 3 missings
            if count == 3:
                if VerificationT0 and VerificationT3 and VerificationT6:
                    data = [T12, T18, T24]
                if VerificationT0 and VerificationT3 and VerificationT12:
                    data = [T6, T18, T24]
                if VerificationT0 and VerificationT3 and VerificationT18:
                    data = [T6, T12, T24]
                if VerificationT0 and VerificationT3 and VerificationT24:
                    data = [T6, T12, T18]
                if VerificationT0 and VerificationT6 and VerificationT12:
                    data = [T3, T18, T24]
                if VerificationT0 and VerificationT6 and VerificationT18:
                    data = [T3, T12, T24]
                if VerificationT0 and VerificationT6 and VerificationT24:
                    data = [T3, T12, T18]
                if VerificationT0 and VerificationT12 and VerificationT18:
                    data = [T3, T6, T24]
                if VerificationT0 and VerificationT12 and VerificationT24:
                    data = [T3, T6, T18]
                if VerificationT0 and VerificationT18 and VerificationT24:
                    data = [T3, T6, T12]
                if VerificationT3 and VerificationT6 and VerificationT12:
                    data = [T0, T18, T24]
                if VerificationT3 and VerificationT6 and VerificationT18:
                    data = [T0, T12, T24]
                if VerificationT3 and VerificationT6 and VerificationT24:
                    data = [T0, T12, T18]
                if VerificationT3 and VerificationT12 and VerificationT18:
                    data = [T0, T6, T24]
                if VerificationT3 and VerificationT12 and VerificationT24:
                    data = [T0, T6, T18]
                if VerificationT3 and VerificationT18 and VerificationT24:
                    data = [T0, T6, T12]
                if VerificationT6 and VerificationT12 and VerificationT18:
                    data = [T0, T3, T24]
                if VerificationT6 and VerificationT12 and VerificationT24:
                    data = [T0, T3, T18]
                if VerificationT6 and VerificationT18 and VerificationT24:
                    data = [T0, T3, T12]
                if VerificationT12 and VerificationT18 and VerificationT24:
                    data = [T0, T3, T6]
            #When there is 2 missings
            if count == 2:
                if VerificationT0 and VerificationT3:
                    data = [T6, T12, T18, T24]
                if VerificationT0 and VerificationT6:
                    data = [T3, T12, T18, T24]
                if VerificationT0 and VerificationT12:
                    data = [T3, T6, T18, T24]
                if VerificationT0 and VerificationT18:
                    data = [T3, T6, T12, T24]
                if VerificationT0 and VerificationT24:
                    data = [T3, T6, T12, T18]
                if VerificationT3 and VerificationT6:
                    data = [T0, T12, T18, T24]
                if VerificationT3 and VerificationT12:
                    data = [T0, T6, T18, T24]
                if VerificationT3 and VerificationT18:
                    data = [T0, T6, T12, T24]
                if VerificationT3 and VerificationT24:
                    data = [T0, T6, T12, T18]
                if VerificationT6 and VerificationT12:
                    data = [T0, T3, T18, T24]
                if VerificationT6 and VerificationT18:
                    data = [T0, T3, T12, T24]
                if VerificationT6 and VerificationT24:
                    data = [T0, T3, T12, T18]
                if VerificationT12 and VerificationT18:
                    data = [T0, T3, T6, T24]
                if VerificationT12 and VerificationT24:
                    data = [T0, T3, T6, T18]
                if VerificationT18 and VerificationT24:
                    data = [T0, T3, T6, T12]
            #When there is 1 missing
            if count == 1:
                if VerificationT0:
                    data = [T3, T6, T12, T18, T24]
                if VerificationT3:
                    data = [T0, T6, T12, T18, T24]
                if VerificationT6:
                    data = [T0, T3, T12, T18, T24]
                if VerificationT12:
                    data = [T0, T3, T6, T18, T24]
                if VerificationT18:
                    data = [T0, T3, T6, T12, T24]
                if VerificationT24:
                    data = [T0, T3, T6, T12, T18]
        else:
            data = [T0, T3, T6, T12, T18, T24]
        #Write in Excel
        if realCol in NominalCol:
            try:
                mode = statistics.mode(data)
            except StatisticsError:
                mode = ""
            worksheet.write(row, atualCol, mode)
            atualCol = atualCol + 1
        else:
            dft = DFT(data)
            dft = dft.tolist()
            dwt = DWT(data)
            none = [""]
            paa_array = []
            r = len(dft)
            if r == 1:
                dft = dft + none + none + none 
            elif r == 2:
                dft = dft + none + none 
            elif r == 3:
                dft = dft + none 
            data = np.asarray(data)
            data_ = pd.DataFrame({"ts1": data})
            size_data = len(data)
            if size_data == 6:
                paa = PAA(data_, 6)
                for l in range(len(paa)):
                    paa_array.append(paa[l].tolist())
                if all_same(data_):
                    sax_array = ['a','a','a','a','a','a']
                else:
                    sax = SAX(data_, 6, "abc")
                    sax_array = []
                    for k in range(len(sax)):
                        sax_array.append(sax[k,0])
                sdl = SDL(data)
                clip = clipping(data)
            elif size_data == 5:
                paa = PAA (data_, 5)
                for l in range(len(paa)):
                    paa_array.append(paa[l].tolist())
                paa_array = paa_array + none
                if all_same(data_):
                    sax_array = ['a','a','a','a','a']
                else:
                    sax = SAX(data_, 5, "abc")
                    sax_array = []
                    for k in range(len(sax)):
                        sax_array.append(sax[k,0])
                
                sax_array = sax_array + none
                sdl = SDL(data)
                sdl = sdl + none
                clip = clipping(data)
                clip = clip + none
            elif size_data == 4:
                paa = PAA (data_, 4)
                for l in range(len(paa)):
                    paa_array.append(paa[l].tolist())
                paa_array = paa_array + none + none 
                if all_same(data_):
                    sax_array = ['a','a','a','a']
                else:
                    sax = SAX(data_, 4, "abc")
                    sax_array = []
                    for k in range(len(sax)):
                        sax_array.append(sax[k,0])
                sax_array = sax_array +  none + none
                sdl = SDL(data)
                sdl = sdl + none + none
                clip = clipping(data)
                clip = clip + none + none
            elif size_data == 3:
                paa = PAA (data_, 3)
                for l in range(len(paa)):
                    paa_array.append(paa[l].tolist())
                paa_array = paa_array + none + none + none 
                if all_same(data_):
                    sax_array = ['a','a','a']
                else:
                    sax = SAX(data_, 3, "abc")
                    sax_array = []
                    for k in range(len(sax)):
                        sax_array.append(sax[k,0])
                sax_array = sax_array + none + none + none
                sdl = SDL(data)
                sdl = sdl + none + none + none
                clip = clipping(data)
                clip = clip + none + none + none
            else:
                paa = PAA (data_, 2)
                for l in range(len(paa)):
                    paa_array.append(paa[l].tolist())
                paa_array = paa_array + none + none + none + none
                if all_same(data_):
                    sax_array = ['a','a']
                else:
                    sax = SAX(data_, 2, "abc")
                    sax_array = []
                    for k in range(len(sax)):
                        sax_array.append(sax[k,0])
                sax_array = sax_array + none + none + none + none
                sdl = SDL(data)
                sdl = sdl + none + none + none + none
                clip = clipping(data)
                clip = clip + none + none + none + none
            dwt_array = []
            for i in range(len(dwt)):
                dwt_array.append(dwt[i].tolist())
            dwt_array = dwt_array[0]
            r = len(dwt_array)
            if r == 1:
                dwt_array = dwt_array + none + none 
            elif r == 2:
                dwt_array = dwt_array + none  
            finaldata = dft + dwt_array + paa_array + sdl + sax_array + clip
            for k in range(len(finaldata)): 
                worksheet.write(row, atualCol + k, str(finaldata[k]))
            atualCol = atualCol + k + 1
    return atualCol

def WriteInformation3m (worksheet, MaxRow, MaxColumn, sheet0):
    
    NominalCol = [39,40,41,42]
    """NotIntegerCol = [9,10,11,12,13,14,15,36,38]"""
    
    # Write pacients information
    for i in range(2,4):
        worksheet.write(0, i-2, sheet0.cell(row = 1, column = i + 1).value)
    
    #Write features Representation 
    indexCol = 2    
    for r in range(4, MaxColumn):
        if r == MaxColumn - 1:
            worksheet.write(0,indexCol, sheet0.cell(row = 1, column = r + 1).value)
        elif r in NominalCol:
            worksheet.write(0,indexCol, sheet0.cell(row = 1, column = r + 1).value + "_MODE")
            indexCol = indexCol + 1
        else:
            worksheet.write(0, indexCol, sheet0.cell(row = 1, column = r + 1).value + "_DFT1")
            worksheet.write(0, indexCol + 1, sheet0.cell(row = 1, column = r + 1).value + "_DFT2")
            worksheet.write(0, indexCol + 2, sheet0.cell(row = 1, column = r + 1).value + "_DWT1")
            worksheet.write(0, indexCol + 3, sheet0.cell(row = 1, column = r + 1).value + "_PAA1")
            worksheet.write(0, indexCol + 4, sheet0.cell(row = 1, column = r + 1).value + "_PAA2")
            worksheet.write(0, indexCol + 5, sheet0.cell(row = 1, column = r + 1).value + "_SDL1")
            worksheet.write(0, indexCol + 6, sheet0.cell(row = 1, column = r + 1).value + "_SAX1")
            worksheet.write(0, indexCol + 7, sheet0.cell(row = 1, column = r + 1).value + "_SAX2")
            worksheet.write(0, indexCol + 8, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING1")
            worksheet.write(0, indexCol + 9, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING2")
            indexCol = indexCol + 10

def WriteInformation6m (worksheet, MaxRow, MaxColumn, sheet0):
    
    NominalCol = [39,40,41,42]
    """NotIntegerCol = [9,10,11,12,13,14,15,36,38]"""
    
    # Write pacients information
    for i in range(2,4):
        worksheet.write(0, i-2, sheet0.cell(row = 1, column = i + 1).value)
    
    #Write features Representation 
    indexCol = 2    
    for r in range(4, MaxColumn):
        if r == MaxColumn - 1:
            worksheet.write(0,indexCol, sheet0.cell(row = 1, column = r + 1).value)
        elif r in NominalCol:
            worksheet.write(0,indexCol, sheet0.cell(row = 1, column = r + 1).value + "_MODE")
            indexCol = indexCol + 1
        else:
            worksheet.write(0, indexCol, sheet0.cell(row = 1, column = r + 1).value + "_DFT1")
            worksheet.write(0, indexCol + 1, sheet0.cell(row = 1, column = r + 1).value + "_DFT2")
            worksheet.write(0, indexCol + 2, sheet0.cell(row = 1, column = r + 1).value + "_DWT1")
            worksheet.write(0, indexCol + 3, sheet0.cell(row = 1, column = r + 1).value + "_DWT2")
            worksheet.write(0, indexCol + 4, sheet0.cell(row = 1, column = r + 1).value + "_PAA1")
            worksheet.write(0, indexCol + 5, sheet0.cell(row = 1, column = r + 1).value + "_PAA2")
            worksheet.write(0, indexCol + 6, sheet0.cell(row = 1, column = r + 1).value + "_PAA3")
            worksheet.write(0, indexCol + 7, sheet0.cell(row = 1, column = r + 1).value + "_SDL1")
            worksheet.write(0, indexCol + 8, sheet0.cell(row = 1, column = r + 1).value + "_SDL2")
            worksheet.write(0, indexCol + 9, sheet0.cell(row = 1, column = r + 1).value + "_SAX1")
            worksheet.write(0, indexCol + 10, sheet0.cell(row = 1, column = r + 1).value + "_SAX2")
            worksheet.write(0, indexCol + 11, sheet0.cell(row = 1, column = r + 1).value + "_SAX3")
            worksheet.write(0, indexCol + 12, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING1")
            worksheet.write(0, indexCol + 13, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING2")
            worksheet.write(0, indexCol + 14, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING3")
            indexCol = indexCol + 15

def WriteInformation12m (worksheet, MaxRow, MaxColumn, sheet0):
    
    NominalCol = [39,40,41,42]
    """NotIntegerCol = [9,10,11,12,13,14,15,36,38]"""
    
    # Write pacients information
    for i in range(2,4):
        worksheet.write(0, i-2, sheet0.cell(row = 1, column = i + 1).value)
    
    #Write features Representation 
    indexCol = 2    
    for r in range(4, MaxColumn):
        if r == MaxColumn - 1:
            worksheet.write(0,indexCol, sheet0.cell(row = 1, column = r + 1).value)
        elif r in NominalCol:
            worksheet.write(0,indexCol, sheet0.cell(row = 1, column = r + 1).value + "_MODE")
            indexCol = indexCol + 1
        else:
            worksheet.write(0, indexCol, sheet0.cell(row = 1, column = r + 1).value + "_DFT1")
            worksheet.write(0, indexCol + 1, sheet0.cell(row = 1, column = r + 1).value + "_DFT2")
            worksheet.write(0, indexCol + 2, sheet0.cell(row = 1, column = r + 1).value + "_DFT3")
            worksheet.write(0, indexCol + 3, sheet0.cell(row = 1, column = r + 1).value + "_DWT1")
            worksheet.write(0, indexCol + 4, sheet0.cell(row = 1, column = r + 1).value + "_DWT2")
            worksheet.write(0, indexCol + 5, sheet0.cell(row = 1, column = r + 1).value + "_PAA1")
            worksheet.write(0, indexCol + 6, sheet0.cell(row = 1, column = r + 1).value + "_PAA2")
            worksheet.write(0, indexCol + 7, sheet0.cell(row = 1, column = r + 1).value + "_PAA3")
            worksheet.write(0, indexCol + 8, sheet0.cell(row = 1, column = r + 1).value + "_PAA4")
            worksheet.write(0, indexCol + 9, sheet0.cell(row = 1, column = r + 1).value + "_SDL1")
            worksheet.write(0, indexCol + 10, sheet0.cell(row = 1, column = r + 1).value + "_SDL2")
            worksheet.write(0, indexCol + 11, sheet0.cell(row = 1, column = r + 1).value + "_SDL3")
            worksheet.write(0, indexCol + 12, sheet0.cell(row = 1, column = r + 1).value + "_SAX1")
            worksheet.write(0, indexCol + 13, sheet0.cell(row = 1, column = r + 1).value + "_SAX2")
            worksheet.write(0, indexCol + 14, sheet0.cell(row = 1, column = r + 1).value + "_SAX3")
            worksheet.write(0, indexCol + 15, sheet0.cell(row = 1, column = r + 1).value + "_SAX4")
            worksheet.write(0, indexCol + 16, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING1")
            worksheet.write(0, indexCol + 17, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING2")
            worksheet.write(0, indexCol + 18, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING3")
            worksheet.write(0, indexCol + 19, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING4")
            indexCol = indexCol + 20

def WriteInformation18m (worksheet, MaxRow, MaxColumn, sheet0):
    
    NominalCol = [39,40,41,42]
    """NotIntegerCol = [9,10,11,12,13,14,15,36,38]"""
    
    # Write pacients information
    for i in range(2,4):
        worksheet.write(0, i-2, sheet0.cell(row = 1, column = i + 1).value)
    
    #Write features Representation 
    indexCol = 2    
    for r in range(4, MaxColumn):
        if r == MaxColumn - 1:
            worksheet.write(0,indexCol, sheet0.cell(row = 1, column = r + 1).value)
        elif r in NominalCol:
            worksheet.write(0,indexCol, sheet0.cell(row = 1, column = r + 1).value + "_MODE")
            indexCol = indexCol + 1
        else:
            worksheet.write(0, indexCol, sheet0.cell(row = 1, column = r + 1).value + "_DFT1")
            worksheet.write(0, indexCol + 1, sheet0.cell(row = 1, column = r + 1).value + "_DFT2")
            worksheet.write(0, indexCol + 2, sheet0.cell(row = 1, column = r + 1).value + "_DFT3")
            worksheet.write(0, indexCol + 3, sheet0.cell(row = 1, column = r + 1).value + "_DWT1")
            worksheet.write(0, indexCol + 4, sheet0.cell(row = 1, column = r + 1).value + "_DWT2")
            worksheet.write(0, indexCol + 5, sheet0.cell(row = 1, column = r + 1).value + "_DWT3")
            worksheet.write(0, indexCol + 6, sheet0.cell(row = 1, column = r + 1).value + "_PAA1")
            worksheet.write(0, indexCol + 7, sheet0.cell(row = 1, column = r + 1).value + "_PAA2")
            worksheet.write(0, indexCol + 8, sheet0.cell(row = 1, column = r + 1).value + "_PAA3")
            worksheet.write(0, indexCol + 9, sheet0.cell(row = 1, column = r + 1).value + "_PAA4")
            worksheet.write(0, indexCol + 10, sheet0.cell(row = 1, column = r + 1).value + "_PAA5")
            worksheet.write(0, indexCol + 11, sheet0.cell(row = 1, column = r + 1).value + "_SDL1")
            worksheet.write(0, indexCol + 12, sheet0.cell(row = 1, column = r + 1).value + "_SDL2")
            worksheet.write(0, indexCol + 13, sheet0.cell(row = 1, column = r + 1).value + "_SDL3")
            worksheet.write(0, indexCol + 14, sheet0.cell(row = 1, column = r + 1).value + "_SDL4")
            worksheet.write(0, indexCol + 15, sheet0.cell(row = 1, column = r + 1).value + "_SAX1")
            worksheet.write(0, indexCol + 16, sheet0.cell(row = 1, column = r + 1).value + "_SAX2")
            worksheet.write(0, indexCol + 17, sheet0.cell(row = 1, column = r + 1).value + "_SAX3")
            worksheet.write(0, indexCol + 18, sheet0.cell(row = 1, column = r + 1).value + "_SAX4")
            worksheet.write(0, indexCol + 19, sheet0.cell(row = 1, column = r + 1).value + "_SAX5")
            worksheet.write(0, indexCol + 20, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING1")
            worksheet.write(0, indexCol + 21, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING2")
            worksheet.write(0, indexCol + 22, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING3")
            worksheet.write(0, indexCol + 23, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING4")
            worksheet.write(0, indexCol + 24, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING5")
            indexCol = indexCol + 25

def WriteInformation24m (worksheet, MaxRow, MaxColumn, sheet0):
    
    NominalCol = [39,40,41,42]
    """NotIntegerCol = [9,10,11,12,13,14,15,36,38]"""
    
    # Write pacients information
    for i in range(2,4):
        worksheet.write(0, i-2, sheet0.cell(row = 1, column = i + 1).value)
    
    #Write features Representation 
    indexCol = 2    
    for r in range(4, MaxColumn):
        if r == MaxColumn - 1:
            worksheet.write(0,indexCol, sheet0.cell(row = 1, column = r + 1).value)
        elif r in NominalCol:
            worksheet.write(0,indexCol, sheet0.cell(row = 1, column = r + 1).value + "_MODE")
            indexCol = indexCol + 1
        else:
            worksheet.write(0, indexCol, sheet0.cell(row = 1, column = r + 1).value + "_DFT1")
            worksheet.write(0, indexCol + 1, sheet0.cell(row = 1, column = r + 1).value + "_DFT2")
            worksheet.write(0, indexCol + 2, sheet0.cell(row = 1, column = r + 1).value + "_DFT3")
            worksheet.write(0, indexCol + 3, sheet0.cell(row = 1, column = r + 1).value + "_DFT4")
            worksheet.write(0, indexCol + 4, sheet0.cell(row = 1, column = r + 1).value + "_DWT1")
            worksheet.write(0, indexCol + 5, sheet0.cell(row = 1, column = r + 1).value + "_DWT2")
            worksheet.write(0, indexCol + 6, sheet0.cell(row = 1, column = r + 1).value + "_DWT3")
            worksheet.write(0, indexCol + 7, sheet0.cell(row = 1, column = r + 1).value + "_PAA1")
            worksheet.write(0, indexCol + 8, sheet0.cell(row = 1, column = r + 1).value + "_PAA2")
            worksheet.write(0, indexCol + 9, sheet0.cell(row = 1, column = r + 1).value + "_PAA3")
            worksheet.write(0, indexCol + 10, sheet0.cell(row = 1, column = r + 1).value + "_PAA4")
            worksheet.write(0, indexCol + 11, sheet0.cell(row = 1, column = r + 1).value + "_PAA5")
            worksheet.write(0, indexCol + 12, sheet0.cell(row = 1, column = r + 1).value + "_PAA6")
            worksheet.write(0, indexCol + 13, sheet0.cell(row = 1, column = r + 1).value + "_SDL1")
            worksheet.write(0, indexCol + 14, sheet0.cell(row = 1, column = r + 1).value + "_SDL2")
            worksheet.write(0, indexCol + 15, sheet0.cell(row = 1, column = r + 1).value + "_SDL3")
            worksheet.write(0, indexCol + 16, sheet0.cell(row = 1, column = r + 1).value + "_SDL4")
            worksheet.write(0, indexCol + 17, sheet0.cell(row = 1, column = r + 1).value + "_SDL5")
            worksheet.write(0, indexCol + 18, sheet0.cell(row = 1, column = r + 1).value + "_SAX1")
            worksheet.write(0, indexCol + 19, sheet0.cell(row = 1, column = r + 1).value + "_SAX2")
            worksheet.write(0, indexCol + 20, sheet0.cell(row = 1, column = r + 1).value + "_SAX3")
            worksheet.write(0, indexCol + 21, sheet0.cell(row = 1, column = r + 1).value + "_SAX4")
            worksheet.write(0, indexCol + 22, sheet0.cell(row = 1, column = r + 1).value + "_SAX5")
            worksheet.write(0, indexCol + 23, sheet0.cell(row = 1, column = r + 1).value + "_SAX6")
            worksheet.write(0, indexCol + 24, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING1")
            worksheet.write(0, indexCol + 25, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING2")
            worksheet.write(0, indexCol + 26, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING3")
            worksheet.write(0, indexCol + 27, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING4")
            worksheet.write(0, indexCol + 28, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING5")
            worksheet.write(0, indexCol + 29, sheet0.cell(row = 1, column = r + 1).value + "_CLIPPING6")
            indexCol = indexCol + 30

def main():
    
    dataFileName = 'new_reuma_mostRecent_withMissingSTATS.xlsx'
    #dataFileName = 'new_reuma.xlsx'
    
    # To read from new_reuma.xlsx file
    wb = openpyxl.load_workbook(dataFileName, data_only = True)
    sheet0 = wb.get_sheet_by_name('0 meses')
    sheet3 = wb.get_sheet_by_name('3 meses')
    sheet6 = wb.get_sheet_by_name('6 meses')
    sheet12 = wb.get_sheet_by_name('12 meses')
    sheet18 = wb.get_sheet_by_name('18 meses')
    sheet24 = wb.get_sheet_by_name('24 meses')
    
    # ==================================================================
    # To write in DataRepresentation.xlsx file
    #workbook = xlsxwriter.Workbook('new_reuma_DataRepresentation.xlsx')
    #worksheet = workbook.add_worksheet('Representation')
    
    MaxColumn = sheet0.max_column
    MaxRow = sheet0.max_row
    
    #WriteInformation(worksheet, MaxRow, MaxColumn, sheet0)
        
    # ==================================================
    # NEW_REUMA --> classFile_
#    for r in range(1, MaxRow):
#        print('Linha: %d' %r)
#        atualCol = 2
#        for k in range(4, MaxColumn):
#            if k == MaxColumn - 1:
#                if estaticas.cell(row = r + 1, column = 4).value == '0m':
#                    worksheet.write(r, atualCol, sheet0.cell(row = r + 1, column = k + 1).value)
#                elif estaticas.cell(row = r + 1, column = 4).value == '3m':
#                    worksheet.write(r, atualCol, sheet3.cell(row = r + 1, column = k + 1).value)
#                elif estaticas.cell(row = r + 1, column = 4).value == '6m':
#                    worksheet.write(r, atualCol, sheet6.cell(row = r + 1, column = k + 1).value)
#                elif estaticas.cell(row = r + 1, column = 4).value == '12m':
#                    worksheet.write(r, atualCol, sheet12.cell(row = r + 1, column = k + 1).value)
#                elif estaticas.cell(row = r + 1, column = 4).value == '18m':
#                    worksheet.write(r, atualCol, sheet18.cell(row = r + 1, column = k + 1).value)
#                else:
#                    worksheet.write(r, atualCol, sheet24.cell(row = r + 1, column = k + 1).value)
#            else:
#                atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
#                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
#                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
#                                                                r, k, atualCol, estaticas.cell(row = r + 1 , column = 4).value, worksheet)
#                worksheet.write(r, 0, sheet0.cell(row = r + 1, column = 3).value)
#                worksheet.write(r, 1, sheet0.cell(row = r + 1, column = 4).value)
#    
#    # EXCEL --> CSV
#    workbook.close()
#    ExcelFile = 'new_reuma_DataRepresentation.xlsx'
#    SheetName = 'Representation'
#    CSVFile = 'new_reuma_DataRepresentation.csv'
#    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT3_classLabelT3
    NewFileName = 'classFile_Representation_T0dataT3_classLabelT3.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation3m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet3.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet3.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '3m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT3_classLabelT3.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT3_classLabelT6
    NewFileName = 'classFile_Representation_T0dataT3_classLabelT6.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation3m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet6.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet6.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '3m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT3_classLabelT6.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT3_classLabelT12
    NewFileName = 'classFile_Representation_T0dataT3_classLabelT12.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation3m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet12.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet12.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '3m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT3_classLabelT12.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT3_classLabelT18
    NewFileName = 'classFile_Representation_T0dataT3_classLabelT18.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation3m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet18.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet18.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '3m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT3_classLabelT18.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT3_classLabelT24
    NewFileName = 'classFile_Representation_T0dataT3_classLabelT24.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation3m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet24.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet24.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '3m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT3_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT6_classLabelT6
    NewFileName = 'classFile_Representation_T0dataT6_classLabelT6.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation6m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet6.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet6.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '6m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT6_classLabelT6.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT6_classLabelT12
    NewFileName = 'classFile_Representation_T0dataT6_classLabelT12.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation6m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet12.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet12.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '6m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT6_classLabelT12.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT6_classLabelT18
    NewFileName = 'classFile_Representation_T0dataT6_classLabelT18.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation6m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet18.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet18.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '6m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT6_classLabelT18.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT6_classLabelT24
    NewFileName = 'classFile_Representation_T0dataT6_classLabelT24.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation6m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet24.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet24.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '6m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT6_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT12_classLabelT12
    NewFileName = 'classFile_Representation_T0dataT12_classLabelT12.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation12m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet12.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet12.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '12m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT12_classLabelT12.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT12classLabelT18
    NewFileName = 'classFile_Representation_T0dataT12_classLabelT18.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation12m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet18.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet18.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '12m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT12_classLabelT18.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT12_classLabelT24
    NewFileName = 'classFile_Representation_T0dataT12_classLabelT24.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation12m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet24.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet24.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '12m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT12_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT18_classLabelT18
    NewFileName = 'classFile_Representation_T0dataT18_classLabelT18.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation18m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet18.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet18.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '18m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT18_classLabelT18.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT18_classLabelT24
    NewFileName = 'classFile_Representation_T0dataT18_classLabelT24.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation18m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet24.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet24.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '18m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT18_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile__T0dataT24_classLabelT24
    NewFileName = 'classFile_Representation_T0dataT24_classLabelT24.xlsx'
    workbook = xlsxwriter.Workbook(NewFileName)
    worksheet = workbook.add_worksheet('Representation')
    WriteInformation24m(worksheet, MaxRow, MaxColumn, sheet0)
    RealRow = 1
    for r in range(1, MaxRow):
        ResponseCode = sheet24.cell(row = r + 1, column = MaxColumn).value
        VerificationResponseCode = (ResponseCode is None) or (ResponseCode is "") or (ResponseCode is ' ')
        if not VerificationResponseCode:
            atualCol = 2
            for k in range(4, MaxColumn):
                if k == MaxColumn - 1:
                    worksheet.write(RealRow, atualCol, sheet24.cell(row = r + 1, column = MaxColumn).value)
                else:
                    atualCol = DataRepresentation(sheet0.cell(row = r + 1, column = k + 1).value, sheet3.cell(row = r + 1, column = k + 1).value, 
                                         sheet6.cell(row = r + 1, column = k + 1).value, sheet12.cell(row = r + 1, column = k + 1).value, 
                                                    sheet18.cell(row = r + 1, column = k + 1).value, sheet24.cell(row = r + 1, column = k + 1).value, 
                                                                RealRow, k, atualCol, '24m', worksheet)
                    worksheet.write(RealRow, 0, sheet0.cell(row = r + 1, column = 3).value)
                    worksheet.write(RealRow, 1, sheet0.cell(row = r + 1, column = 4).value)
            RealRow += 1
    # EXCEL --> CSV
    workbook.close()
    ExcelFile = NewFileName
    SheetName = 'Representation'
    CSVFile = 'classFile_Representation_T0dataT24_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
   
if __name__ == "__main__":
    main()
