# Reuma Classification
# Files Generator from a given file of type [new_reuma.xlsx], for classification

import xlsxwriter
import openpyxl
import xlrd
import unicodecsv

def createNewFileForClassification(onUse1, onUse2, NewFileName, checkT24T24):
    
    # To write in classFile.xlsx file
    workbook = xlsxwriter.Workbook(NewFileName, {'constant_memory': True})
    worksheet = workbook.add_worksheet('T0')
    
    # Without possible explanation when onUse1 is sheet24 it counts one extra column????
    # This cheating fixes it...
    #if (checkT24T24):
    #    onUse1MaxColumn = onUse1.max_column-1
    #else:
    onUse1MaxColumn = onUse1.max_column
        
    sheet0MaxColumn = sheet0.max_column
    onUse1MaxRow = onUse1.max_row
    
    # Write attributes' names in the new file (first line)
    indexCol = 0
    for k in range(3, onUse1MaxColumn + 1):
        # =============== ADD DATA FROM T0 ==================================
        # Write attributes' names of data T0 in the new file (first line)
        # Write all except cod_resposta_das
        if str(onUse1.cell(row = 1, column = k).value).find("cod_resposta_das") < 0:
            worksheet.write(0, indexCol, sheet0.cell(row = 1, column = k).value + "_T0")
            worksheet.write(0, indexCol+sheet0MaxColumn-4, onUse1.cell(row = 1, column = k).value)
            indexCol = indexCol+1
        # Save index with response code (label)
        if str(onUse1.cell(row = 1, column = k).value).find("cod_resposta_das_label") >= 0:
            labelCol = k
    
    # Add class label (cod_resposta_das_TX)
    worksheet.write(0, indexCol+sheet0MaxColumn-4, "cod_resposta_das_TX")
    
    # Write attributes values in the new file
    # Response code corresponds to the 24m while data corresponds to 0m
    # i = iterator over rows
    # j = iterator over columns
    itRow = 1
    for i in range(2, onUse1MaxRow + 1):
        if str(onUse2.cell(row = i, column = labelCol).value).find("C0") >= 0 or str(onUse2.cell(row = i, column = labelCol).value).find("C1") >= 0 or str(onUse2.cell(row = i, column = labelCol).value).find("C2") >= 0:
            indexCol = 0
            for j in range(3, onUse1MaxColumn + 1):
                if str(onUse1.cell(row = 1, column = j).value).find("cod_resposta_das") < 0:
                    # Add T0 data
                    worksheet.write(itRow, indexCol, sheet0.cell(row = i, column = j).value)
                    # Add onUse1 information
                    worksheet.write(itRow, indexCol+sheet0MaxColumn-4, onUse1.cell(row = i, column = j).value)
                    indexCol = indexCol+1
            # Add label from onUse2
            worksheet.write(itRow, indexCol+sheet0MaxColumn-4, onUse2.cell(row = i, column = labelCol).value)  
            itRow = itRow+1

def convertXLS2CSV(ExcelFile, SheetName, CSVFile):
    # Converts an Excel file to a CSV file.
    # If the excel file has multiple worksheets, only the first worksheet is converted.
    # Uses unicodecsv, so it will handle Unicode characters.
    # Uses a recent version of xlrd, so it should handle old .xls and new .xlsx equally well.
    

    wb = xlrd.open_workbook(ExcelFile)
    sh = wb.sheet_by_name(SheetName)

    fh = open(CSVFile,"wb")
    csv_out = unicodecsv.writer(fh, encoding='utf-8')

    for row_number in range (sh.nrows):
        csv_out.writerow(sh.row_values(row_number))

    fh.close()


# Main
if __name__ == "__main__":
    
    # ================= CHOOSE WHICH NEW_REUMA.XLSX =======================
    dataFileName = 'new_Reuma_ImputatedMissingValuesSTATS.xlsx'
    # dataFileName = 'new_reuma.xlsx'
    # =====================================================================
    
    # To read from new_reuma.xlsx file
    wb = openpyxl.load_workbook(dataFileName, data_only = True)
    sheet = wb.get_sheet_by_name('Estaticas')
    sheet0 = wb.get_sheet_by_name('0 meses')
    sheet3 = wb.get_sheet_by_name('3 meses')
    sheet6 = wb.get_sheet_by_name('6 meses')
    sheet12 = wb.get_sheet_by_name('12 meses')
    sheet18 = wb.get_sheet_by_name('18 meses')
    sheet24 = wb.get_sheet_by_name('24 meses')
    
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT0_classLabelT0
    onUse1 = sheet0 # data
    onUse2 = sheet0 # label
    NewFileName = 'classFile_T0dataT0_classLabelT0.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT0_classLabelT0.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT0_classLabelT3
    onUse1 = sheet0 # data
    onUse2 = sheet3 # label
    NewFileName = 'classFile_T0dataT0_classLabelT3.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT0_classLabelT3.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT0_classLabelT6
    onUse1 = sheet0 # data
    onUse2 = sheet6 # label
    NewFileName = 'classFile_T0dataT0_classLabelT6.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT0_classLabelT6.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT0_classLabelT12
    onUse1 = sheet0 # data
    onUse2 = sheet12 # label
    NewFileName = 'classFile_T0dataT0_classLabelT12.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT0_classLabelT12.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT0_classLabelT18
    onUse1 = sheet0 # data
    onUse2 = sheet18 # label
    NewFileName = 'classFile_T0dataT0_classLabelT18.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT0_classLabelT18.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT0_classLabelT24
    onUse1 = sheet0 # data
    onUse2 = sheet24 # label
    NewFileName = 'classFile_T0dataT0_classLabelT24.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT0_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT3_classLabelT3
    onUse1 = sheet3 # data
    onUse2 = sheet3 # label
    NewFileName = 'classFile_T0dataT3_classLabelT3.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT3_classLabelT3.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT3_classLabelT6
    onUse1 = sheet3 # data
    onUse2 = sheet6 # label
    NewFileName = 'classFile_T0dataT3_classLabelT6.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT3_classLabelT6.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT3_classLabelT12
    onUse1 = sheet3 # data
    onUse2 = sheet12 # label
    NewFileName = 'classFile_T0dataT3_classLabelT12.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT3_classLabelT12.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT3_classLabelT18
    onUse1 = sheet3 # data
    onUse2 = sheet18 # label
    NewFileName = 'classFile_T0dataT3_classLabelT18.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT3_classLabelT18.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT3_classLabelT24
    onUse1 = sheet3 # data
    onUse2 = sheet24 # label
    NewFileName = 'classFile_T0dataT3_classLabelT24.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT3_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT6_classLabelT6
    onUse1 = sheet6 # data
    onUse2 = sheet6 # label
    NewFileName = 'classFile_T0dataT6_classLabelT6.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT6_classLabelT6.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT6_classLabelT12
    onUse1 = sheet6 # data
    onUse2 = sheet12 # label
    NewFileName = 'classFile_T0dataT6_classLabelT12.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT6_classLabelT12.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT6_classLabelT18
    onUse1 = sheet6 # data
    onUse2 = sheet18 # label
    NewFileName = 'classFile_T0dataT6_classLabelT18.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT6_classLabelT18.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT6_classLabelT24
    onUse1 = sheet6 # data
    onUse2 = sheet24 # label
    NewFileName = 'classFile_T0dataT6_classLabelT24.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT6_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT12_classLabelT12
    onUse1 = sheet12 # data
    onUse2 = sheet12 # label
    NewFileName = 'classFile_T0dataT12_classLabelT12.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT12_classLabelT12.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT12_classLabelT18
    onUse1 = sheet12 # data
    onUse2 = sheet18 # label
    NewFileName = 'classFile_T0dataT12_classLabelT18.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT12_classLabelT18.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT12_classLabelT24
    onUse1 = sheet12 # data
    onUse2 = sheet24 # label
    NewFileName = 'classFile_T0dataT12_classLabelT24.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT12_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT18_classLabelT18
    onUse1 = sheet18 # data
    onUse2 = sheet18 # label
    NewFileName = 'classFile_T0dataT18_classLabelT18.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT18_classLabelT18.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT18_classLabelT24
    onUse1 = sheet18 # data
    onUse2 = sheet24 # label
    NewFileName = 'classFile_T0dataT18_classLabelT24.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, False)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT18_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    # ==================================================
    # NEW_REUMA --> classFile_T0dataT24_classLabelT24
    onUse1 = sheet24 # data
    onUse2 = sheet24 # label
    NewFileName = 'classFile_T0dataT24_classLabelT24.xlsx'
    createNewFileForClassification(onUse1, onUse2, NewFileName, True)
    # EXCEL --> CSV
    ExcelFile = NewFileName
    SheetName = 'T0'
    CSVFile = 'classFile_T0dataT24_classLabelT24.csv'
    convertXLS2CSV(ExcelFile, SheetName, CSVFile)
    